#populate_excel.py
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import os
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, filename='populate_excel.log', filemode='a',
                    format='%(name)s - %(levelname)s - %(message)s')

def populate_template(data, template_path, output_path):
    try:
        logging.info("Starting to populate template")
        
        # Load the Excel template
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active

        # Insert and resize the image
        img_path = os.path.join(os.path.dirname(__file__), 'eahead.jpg')
        img = Image(img_path)
        img.width, img.height = img.width * 0.43, img.height * 0.60
        sheet.add_image(img, 'A1')
        logging.debug(f"Image added at path: {img_path}")

        # Common setup for cell alignment
        center_aligned_text = Alignment(horizontal='center', wrapText=True)

        # Populate cells with provided data
        cells_to_populate = {
            'B5': data['school'],
            'H4': datetime.strptime(data['period_ending'], '%Y-%m-%d').date(),
            'H5': data['trip_purpose'],
            'B4': data['employee_department']
        }

        for cell_ref, value in cells_to_populate.items():
            cell = sheet[cell_ref]
            cell.alignment = center_aligned_text
            cell.value = value
            logging.debug(f"Populated {cell_ref} with value: {value}")

        # Handling date names and values
        populate_dates(sheet, data['period_ending'])

        # Handling travel dates for per diem fields
        if 'yes' == data.get('travel') and all(data.get(key) for key in ['travel_start_date', 'travel_end_date']):
            populate_travel_dates(sheet, data['travel_start_date'], data['travel_end_date'])

        # Save the populated template to a new file
        wb.save(output_path)
        logging.info(f"Template populated and saved to: {output_path}")

    except Exception as e:
        logging.error(f"An error occurred while populating the template: {e}", exc_info=True)
        raise

def populate_dates(sheet, period_ending):
    try:
        period_ending_date = datetime.strptime(period_ending, '%Y-%m-%d')
        day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        for i, day_name in enumerate(day_names):
            day_cell = sheet.cell(row=7, column=4+i)
            day_cell.value = f"Date\n{day_name}"
            day_cell.alignment = Alignment(horizontal='center', wrapText=True)
            logging.debug(f"Populated day name: {day_name} at row 7, column {4+i}")

        for i in range(6, -1, -1):
            date_cell = sheet.cell(row=8, column=4+i)
            date_cell.value = (period_ending_date - timedelta(days=6-i)).date()
            date_cell.alignment = Alignment(horizontal='center', wrapText=True)
            logging.debug(f"Populated date: {date_cell.value} at row 8, column {4+i}")
    except Exception as e:
        logging.error(f"An error occurred while populating dates: {e}", exc_info=True)
        raise

def populate_travel_dates(sheet, travel_start_date_str, travel_end_date_str):
    try:
        travel_start_date = datetime.strptime(travel_start_date_str, '%Y-%m-%d')
        travel_end_date = datetime.strptime(travel_end_date_str, '%Y-%m-%d')
        
        for i in range(4, 11):
            cell_date = sheet.cell(row=8, column=i).value
            if cell_date and travel_start_date.date() <= cell_date <= travel_end_date.date():
                if cell_date == travel_start_date.date():
                    sheet.cell(row=21, column=i).value = 30.00  # Only dinner on start date
                elif cell_date == travel_end_date.date():
                    sheet.cell(row=19, column=i).value = 5.00  # Only breakfast on end date
                else:
                    sheet.cell(row=19, column=i).value = 5.00  # Breakfast
                    sheet.cell(row=21, column=i).value = 30.00  # Dinner
                logging.debug(f"Populated per diem for date: {cell_date} at column {i}")
    except Exception as e:
        logging.error(f"An error occurred while populating travel dates: {e}", exc_info=True)
        raise